from openpyxl import load_workbook

MIN_ROW = 7
FORTNIGHTLY_MAX_ROW = 697
MONTHLY_MAX_ROW = 534

# MONTHLY_WRITING_COLS : 16, 17, 18
# FORTNIGHTLY_WRITING_COLS : 10, 11, 12


def if_valid_data_row(row_0):
    if row_0 == 1:
        return False
    else:
        try:
            row_0 + 0
        except:
            return False
        else:
            return True


def read_from_file():
    print('Start loading...')
    wb = load_workbook(filename='fortnightly.xlsx', read_only=True)
    print('Loading finished')

    results = []

    for row in wb['Table 1'].iter_rows(min_row=MIN_ROW, max_row=FORTNIGHTLY_MAX_ROW, values_only=True):
        if if_valid_data_row(row[0]):
            result = []
            for cell in row:
                if cell == '—':
                    cell = 0
                if cell is not None:
                    result.append(cell)

            result_split = []
            for idx in range(0, 5):
                result_split.append(result[idx*3: idx*3+3])

            results += result_split

    print("Sorting results...")
    sorted_results = sorted(results, key=lambda res: res[0])
    print(sorted_results)
    return sorted_results


def write_to_file(results):
    print('Start writing...')
    wb = load_workbook('result.xlsx')
    ws = wb['Info']
    for idx, result in enumerate(results):
        ws.cell(row=idx + 2, column=10).value = result[0]
        ws.cell(row=idx + 2, column=11).value = result[1]
        ws.cell(row=idx + 2, column=12).value = result[2]
    wb.save(filename='result.xlsx')
    print("Finished")


def main():

    write_to_file(read_from_file())


if __name__ == "__main__":
    # execute only if run as a script
    main()
